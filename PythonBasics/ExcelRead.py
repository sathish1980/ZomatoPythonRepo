import openpyxl as openpyxl
class ExcelRead:
    Dict = {}
    def fileread(self):
        book = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\WeekdayPython\\inputFile\\InputFile.xlsx")
        sheet = book["Login"]
        cell = sheet.cell(row=1, column=2) # move in to the specific cell
        print(cell)
        # sheet.cell(row=5, column=2).value = "newvalue"  # write the data in to the excel sheet
        # for mar row
        max_row = sheet.max_row
        print(max_row)
        # for mar column
        max_column = sheet.max_column
        print(max_column)
        for i in range(2, max_row + 1):
            for j in range(1, max_column + 1):
                #cell = sheet.cell(row=i, column=j)
                #print(cell.value)
                self.Dict[(sheet.cell(row=1, column=j).value) + str(i-1)] = sheet.cell(row=i, column=j).value

        print(self.Dict)

obj=ExcelRead()
obj.fileread()