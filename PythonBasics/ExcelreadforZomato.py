import openpyxl as openpyxl
class ExcelReadForZomato:
    Dict = {}
    def filereadforZomato(self,sheetName,HotelName):
        book = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\WeekdayPython\\inputFile\\InputFile.xlsx")
        sheet = book[sheetName]
        #print(sheetName)
        cell = sheet.cell(row=1, column=2) # move in to the specific cell
        #print(cell)
        # sheet.cell(row=5, column=2).value = "newvalue"  # write the data in to the excel sheet
        # for mar row
        max_row = sheet.max_row
        #print(max_row)
        # for max column
        max_columnv = sheet.max_column
        #print(max_columnv)
        for j in range(1, max_columnv+1):
                cell = sheet.cell(row=1, column=j)
                columnvalue=sheet.cell(row=1, column=j).value
                if columnvalue==HotelName:
                    column_Value=j
                    break



        for i in range(2, max_row + 1):
            cell = sheet.cell(row=i, column=column_Value)
            #print(cell.value)
            self.Dict[(sheet.cell(row=1, column=j).value) + str(i-1)] = str(sheet.cell(row=i, column=j).value)

        #print(self.Dict)
        return self.Dict

    def filereadmaxrow(self,sheetName):
        book = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\WeekdayPython\\inputFile\\InputFile.xlsx")
        sheet = book[sheetName]
        max_row = sheet.max_row
        return max_row
